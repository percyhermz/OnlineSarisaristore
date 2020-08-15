from e_store.models import Product
from django.core.files import File
from openpyxl import load_workbook
from PIL import Image
import io
import os



os.environ['DJANGO_SETTINGS_MODULE'] = 'settings'
wb_url = r'./products to add/sampleproducts.xlsx'
image_url = r"./products to add"

# opens an image file using django File object
def imageFileOpen(image_name):
    image  = imageBoxResize(os.path.join(image_url, image_name))
    f = File(image)
    return f

# converter of text category from object to choice class
def category_to_choices(category): 
    switcher = { 
        'can': Product.ProductCategory.CANNED,
        'nood': Product.ProductCategory.NOODLES, 
        'veg': Product.ProductCategory.VEGGIES,
        'rice': Product.ProductCategory.RICE,
    }
    return switcher.get(category, Product.ProductCategory.CANNED) 

# Creates new object with a dictionary argument
def createNewObject(product_info_list):
    obj = Product(name=product_info_list['title'],
                        code=product_info_list['code'],
                        price=product_info_list['price'],
                        category=category_to_choices(product_info_list['category']))

    obj.save()
    index = 0

    for image_name in product_info_list['image']:
        print("relative path of file: " +image_name)
        print("the path that you are saving " + os.path.join(image_url, image_name))
        index += 1
        obj.images.create(image=imageFileOpen(image_name), num=index)
        print(obj.images.all())
    
    obj.save()

# Starts the program, takes an url of EXCEL file then extract data each row to convert into ProductModels
def loadwb(wb_url):
    wb = load_workbook(wb_url)
    wb.active = wb["Products"]
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=5, max_row=19, values_only=True):
        n_row = []

        for any_cell in row:
            n_row.append(any_cell)

        obj_to_add = {'title': n_row[0],
                    'price': n_row[1],
                    'code': n_row[2],
                    'image': n_row[3].split(", "),
                    'category': n_row[4],
                    }
        createNewObject(obj_to_add)
        print(obj_to_add)
        
    wb.close()
    print("FINISHED UPLOADING PRODUCT DATA")

# Resize Image into a box, takes an image url and size of box dimension
def imageBoxResize(image_url, size=500):
    image  = Image.open(image_url)
    name = image.filename
    converted = image.convert('RGB')
    converted.filename = name
    print('Now editing file : ' + converted.filename)
    if (image.height > image.width) :
        print("height is bigger than width, proceeding to square the photo to "+str(size)+"x"+str(size)+".")
        offset_x = converted.width % 100
        offset_y = converted.height - (image.width - offset_x)
        offset_halves = [offset_x//2, offset_y//2]
        new_image = converted.crop((offset_halves[0],
                                offset_halves[1],
                                converted.width - (offset_x - offset_halves[0]),
                                converted.height-(offset_y - offset_halves[1])))
        print("Resizing to "+str(size)+"x"+str(size)+" box dimensions.")
        new_image.thumbnail((size, size))
        thumb_io = io.BytesIO()
        new_image.save(thumb_io, format='JPEG')
        thumb_io.name = os.path.basename(converted.filename)
    elif (image.height < image.width ):
        print("Width is bigger than height, proceeding to square the photo to "+str(size)+"x"+str(size)+".")
        offset_y = converted.height % 100
        offset_x = converted.width - (converted.height - offset_y)
        offset_halves = [offset_x//2, offset_y//2]
        new_image = converted.crop((offset_halves[0],
                                offset_halves[1],
                                converted.width - (offset_x - offset_halves[0]),
                                converted.height-(offset_y - offset_halves[1])))
        print("Resizing to "+str(size)+"x"+str(size)+" box dimensions.")
        new_image.thumbnail((size, size))
        thumb_io = io.BytesIO()
        new_image.save(thumb_io, format='JPEG')
        thumb_io.name = os.path.basename(converted.filename)
    else:
        print("Dimensions are equal, proceeding to resize photo to "+str(size)+"x"+str(size)+".")
        new_image = converted.copy()
        new_image.thumbnail((size,size))
        thumb_io = io.BytesIO()
        new_image.save(thumb_io, format='JPEG')
        thumb_io.name = os.path.basename(converted.filename)

    return  thumb_io
        
def start():
    loadwb(wb_url)
    print('Finished product maker process!')



